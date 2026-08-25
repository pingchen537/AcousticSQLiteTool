"""Generate a formatted multi-sheet offline Excel acoustic report.

Requires openpyxl, which is already present in the user's Python environment.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Union

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as error:  # pragma: no cover - executed on the production PC
    raise ImportError(
        "產生 Excel 需要 openpyxl。請在有網路的電腦準備 openpyxl wheel，"
        "再離線安裝到產線環境。"
    ) from error

from html_report import (
    CPK_COLUMNS,
    LABELS,
    MEASUREMENT_COLUMNS,
    RETEST_COLUMNS,
    SPEC_COLUMNS,
    SUMMARY_COLUMNS,
    TEST_COLUMNS,
)


PathLike = Union[str, Path]
NAVY = "12304A"
BLUE = "2563EB"
GREEN = "15803D"
LIGHT_GREEN = "DCFCE7"
RED = "B91C1C"
LIGHT_RED = "FEE2E2"
AMBER = "B45309"
LIGHT_AMBER = "FEF3C7"
LIGHT_BLUE = "E8F1FB"
LIGHT_GRAY = "F4F7FA"
WHITE = "FFFFFF"
GRID = "D9E2EA"
TEXT = "1F2937"
MUTED = "64748B"
BODY_FONT = "Microsoft JhengHei"
THIN_GRAY = Side(style="thin", color=GRID)


def _display_label(key: str) -> str:
    return LABELS.get(key, key)


def _excel_value(value: Any, key: str) -> Any:
    if value is None:
        return None
    if key in ("anchor_time", "start_time", "first_test_time", "latest_test_time"):
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            return str(value)
    if key in ("effective_from", "effective_to"):
        try:
            return date.fromisoformat(str(value))
        except ValueError:
            return str(value)
    return value


def _format_filter_text(metadata: Dict[str, Any]) -> str:
    filters = metadata.get("filters", {})
    names = {
        "project_code": "機種",
        "station": "測站",
        "sn": "SN",
        "date_from": "起始日期",
        "date_to": "結束日期",
    }
    values = [
        "{}={}".format(names[key], filters[key])
        for key in ("project_code", "station", "sn", "date_from", "date_to")
        if filters.get(key)
    ]
    return "；".join(values) if values else "全部資料"


def _style_title(worksheet: Any, title: str, subtitle: str, last_column: int) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = worksheet.cell(1, 1, title)
    title_cell.font = Font(name=BODY_FONT, size=18, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 32

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    subtitle_cell = worksheet.cell(2, 1, subtitle)
    subtitle_cell.font = Font(name=BODY_FONT, size=9, color=MUTED)
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[2].height = 30


def _style_header(cell: Any) -> None:
    cell.font = Font(name=BODY_FONT, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="medium", color=BLUE))


def _apply_cell_format(cell: Any, key: str) -> None:
    cell.font = Font(name=BODY_FONT, size=9, color=TEXT)
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(bottom=THIN_GRAY)
    if key in ("anchor_time", "start_time", "first_test_time", "latest_test_time"):
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    elif key in ("effective_from", "effective_to"):
        cell.number_format = "yyyy-mm-dd"
    elif key.endswith("_pct"):
        cell.number_format = '0.00"%"'
    elif key in (
        "duration_sec",
        "target",
        "lsl",
        "usl",
        "measured_value",
        "mean",
        "sample_stddev",
        "cpl",
        "cpu",
        "cpk",
    ):
        cell.number_format = "0.0000"
    elif key in (
        "total_attempts",
        "total_sn",
        "pass_attempts",
        "fail_attempts",
        "unknown_attempts",
        "first_pass_sn",
        "final_pass_sn",
        "retested_sn",
        "attempt_no",
        "attempt_count",
        "n",
    ):
        cell.number_format = "0"
    if key in ("source_note", "error_message"):
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _autofit(worksheet: Any, columns: Sequence[str], start_row: int, end_row: int) -> None:
    for column_index, key in enumerate(columns, start=1):
        maximum = len(_display_label(key))
        for row_index in range(start_row, min(end_row, start_row + 300) + 1):
            value = worksheet.cell(row_index, column_index).value
            maximum = max(maximum, len(str(value)) if value is not None else 0)
        cap = 46 if key in ("source_note", "error_message") else 28
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            cap, max(10, maximum + 2)
        )


def _add_table(
    worksheet: Any,
    rows: Sequence[Dict[str, Any]],
    columns: Sequence[str],
    table_name: str,
    title: str,
    metadata: Dict[str, Any],
) -> None:
    last_column = max(1, len(columns))
    subtitle = "產生時間：{}｜篩選：{}".format(
        metadata.get("generated_at", ""), _format_filter_text(metadata)
    )
    _style_title(worksheet, title, subtitle, last_column)
    header_row = 4
    for column_index, key in enumerate(columns, start=1):
        cell = worksheet.cell(header_row, column_index, _display_label(key))
        _style_header(cell)
    worksheet.row_dimensions[header_row].height = 32

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, key in enumerate(columns, start=1):
            cell = worksheet.cell(row_index, column_index, _excel_value(row.get(key), key))
            _apply_cell_format(cell, key)
        if row_index % 2 == 0:
            for column_index in range(1, last_column + 1):
                worksheet.cell(row_index, column_index).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    end_row = header_row + max(1, len(rows))
    if rows:
        reference = "A{}:{}{}".format(
            header_row,
            get_column_letter(last_column),
            end_row,
        )
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.cell(header_row + 1, 1, "沒有符合資料")
        worksheet.cell(header_row + 1, 1).font = Font(name=BODY_FONT, italic=True, color=MUTED)

    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = "A{}:{}{}".format(
        header_row,
        get_column_letter(last_column),
        max(header_row, end_row),
    )
    worksheet.print_title_rows = "1:4"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddFooter.center.text = "Page &P / &N"
    worksheet.oddFooter.right.text = "Offline Acoustic Report"
    _autofit(worksheet, columns, header_row, end_row)

    result_columns = {
        key: index + 1
        for index, key in enumerate(columns)
        if key in ("overall_result", "evaluated_result", "first_result", "latest_result")
    }
    for column_index in result_columns.values():
        letter = get_column_letter(column_index)
        if rows:
            data_range = "{}5:{}{}".format(letter, letter, end_row)
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=['{}5="FAIL"'.format(letter)],
                    fill=PatternFill("solid", fgColor=LIGHT_RED),
                    font=Font(name=BODY_FONT, bold=True, color=RED),
                ),
            )
            worksheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=['{}5="PASS"'.format(letter)],
                    fill=PatternFill("solid", fgColor=LIGHT_GREEN),
                    font=Font(name=BODY_FONT, bold=True, color=GREEN),
                ),
            )

    if "cpk" in columns and rows:
        cpk_column = columns.index("cpk") + 1
        cpk_letter = get_column_letter(cpk_column)
        cpk_range = "{}5:{}{}".format(cpk_letter, cpk_letter, end_row)
        worksheet.conditional_formatting.add(
            cpk_range,
            CellIsRule(
                operator="lessThan",
                formula=["1.33"],
                fill=PatternFill("solid", fgColor=LIGHT_RED),
                font=Font(name=BODY_FONT, bold=True, color=RED),
            ),
        )
        worksheet.conditional_formatting.add(
            cpk_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["1.33"],
                fill=PatternFill("solid", fgColor=LIGHT_GREEN),
                font=Font(name=BODY_FONT, bold=True, color=GREEN),
            ),
        )


def _summary_totals(summary: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    total_sn = sum(int(row.get("total_sn") or 0) for row in summary)
    retested_sn = sum(int(row.get("retested_sn") or 0) for row in summary)
    final_pass_sn = sum(int(row.get("final_pass_sn") or 0) for row in summary)
    return {
        "total_attempts": sum(int(row.get("total_attempts") or 0) for row in summary),
        "total_sn": total_sn,
        "pass_attempts": sum(int(row.get("pass_attempts") or 0) for row in summary),
        "fail_attempts": sum(int(row.get("fail_attempts") or 0) for row in summary),
        "final_yield_pct": 100.0 * final_pass_sn / total_sn if total_sn else None,
        "retest_rate_pct": 100.0 * retested_sn / total_sn if total_sn else None,
    }


def _build_summary_sheet(workbook: Workbook, report_data: Dict[str, Any]) -> None:
    worksheet = workbook.create_sheet("Summary")
    metadata = report_data["metadata"]
    _style_title(
        worksheet,
        metadata["title"],
        "產生時間：{}｜篩選：{}｜完全離線".format(
            metadata["generated_at"], _format_filter_text(metadata)
        ),
        13,
    )
    totals = _summary_totals(report_data["summary"])
    cards = (
        ("測試次數", totals["total_attempts"], BLUE),
        ("不重複 SN", totals["total_sn"], BLUE),
        ("PASS 次數", totals["pass_attempts"], GREEN),
        ("FAIL 次數", totals["fail_attempts"], RED),
        ("最終良率(%)", totals["final_yield_pct"], GREEN),
        ("重測率(%)", totals["retest_rate_pct"], AMBER),
    )
    for index, (label, value, color) in enumerate(cards):
        start_column = 1 + index * 2
        worksheet.merge_cells(
            start_row=4,
            start_column=start_column,
            end_row=4,
            end_column=start_column + 1,
        )
        worksheet.merge_cells(
            start_row=5,
            start_column=start_column,
            end_row=6,
            end_column=start_column + 1,
        )
        label_cell = worksheet.cell(4, start_column, label)
        value_cell = worksheet.cell(5, start_column, value)
        label_cell.font = Font(name=BODY_FONT, size=9, color=MUTED)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.font = Font(name=BODY_FONT, size=18, bold=True, color=color)
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if label.endswith("(%)"):
            value_cell.number_format = '0.00"%"'
        for row_index in (4, 5, 6):
            for column_index in range(start_column, start_column + 2):
                worksheet.cell(row_index, column_index).border = Border(
                    left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
                )

    start_row = 9
    for column_index, key in enumerate(SUMMARY_COLUMNS, start=1):
        _style_header(worksheet.cell(start_row, column_index, _display_label(key)))
    for row_index, row in enumerate(report_data["summary"], start=start_row + 1):
        for column_index, key in enumerate(SUMMARY_COLUMNS, start=1):
            cell = worksheet.cell(row_index, column_index, _excel_value(row.get(key), key))
            _apply_cell_format(cell, key)

    if report_data["summary"]:
        chart_data_start = start_row
        pass_column = SUMMARY_COLUMNS.index("pass_attempts") + 1
        fail_column = SUMMARY_COLUMNS.index("fail_attempts") + 1
        project_column = SUMMARY_COLUMNS.index("project_code") + 1
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "各機種 PASS / FAIL 次數"
        chart.y_axis.title = "測試次數"
        chart.x_axis.title = "機種"
        chart.height = 8
        chart.width = 15
        chart_data = Reference(
            worksheet,
            min_col=pass_column,
            max_col=fail_column,
            min_row=chart_data_start,
            max_row=start_row + len(report_data["summary"]),
        )
        categories = Reference(
            worksheet,
            min_col=project_column,
            min_row=start_row + 1,
            max_row=start_row + len(report_data["summary"]),
        )
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend.position = "b"
        worksheet.add_chart(chart, "A14")

    warnings = metadata.get("warnings") or []
    if warnings:
        warning_row = max(31, start_row + len(report_data["summary"]) + 3)
        worksheet.merge_cells(
            start_row=warning_row,
            start_column=1,
            end_row=warning_row + len(warnings),
            end_column=13,
        )
        cell = worksheet.cell(
            warning_row,
            1,
            "注意事項\n" + "\n".join("• " + item for item in warnings),
        )
        cell.fill = PatternFill("solid", fgColor=LIGHT_AMBER)
        cell.font = Font(name=BODY_FONT, size=9, color=AMBER)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index in range(1, 14):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 14
    worksheet.freeze_panes = "A9"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.oddFooter.center.text = "Page &P / &N"


def generate_excel_report(report_data: Dict[str, Any], output_path: PathLike) -> Path:
    """Create the complete multi-sheet XLSX report."""
    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = report_data["metadata"]["title"]
    workbook.properties.subject = "ST01 Acoustic Production Report"
    workbook.properties.creator = "AcousticSQLiteTool"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    _build_summary_sheet(workbook, report_data)
    sheet_definitions = (
        ("All_Tests", report_data["tests"], TEST_COLUMNS, "AllTestsTable", "全部測試"),
        ("PASS", report_data["pass_tests"], TEST_COLUMNS, "PassTestsTable", "PASS 測試"),
        ("FAIL", report_data["fail_tests"], TEST_COLUMNS, "FailTestsTable", "FAIL 測試"),
        ("Retest", report_data["retests"], RETEST_COLUMNS, "RetestTable", "重測 SN"),
        (
            "CPK_First_Valid",
            report_data["cpk_first"],
            CPK_COLUMNS,
            "CpkFirstTable",
            "每個 SN 第一次有效量測 CPK",
        ),
        (
            "CPK_All_Attempts",
            report_data["cpk_all"],
            CPK_COLUMNS,
            "CpkAllTable",
            "所有測試（包含重測）CPK",
        ),
        (
            "Measurements",
            report_data["measurements"],
            MEASUREMENT_COLUMNS,
            "MeasurementsTable",
            "量測值與規格判定",
        ),
        (
            "Specifications",
            report_data["specifications"],
            SPEC_COLUMNS,
            "SpecificationsTable",
            "規格版本",
        ),
    )
    for sheet_name, rows, columns, table_name, title in sheet_definitions:
        worksheet = workbook.create_sheet(sheet_name)
        _add_table(
            worksheet,
            rows,
            columns,
            table_name,
            title,
            report_data["metadata"],
        )

    workbook.active = 0
    workbook.save(str(path))
    return path
