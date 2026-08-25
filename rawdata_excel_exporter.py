"""Export daily RawData-style Excel workbooks from acoustic SQLite.

Example
-------
python rawdata_excel_exporter.py --date 2026-07-02 --measurement FREQRESP_IN_1

The first worksheet matches the requested merged layout:
project, SN, TestDate, TestTime, Result, followed by measurement values.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from measurement_data_service import (
    build_export_frame,
    connect_readonly,
    list_export_groups,
    normalize_measurement_name,
    query_current_specs,
    raw_measurement_name,
    validate_iso_date,
)


SCRIPT_DIRECTORY = Path(__file__).resolve().parent
DEFAULT_DATABASE = SCRIPT_DIRECTORY / "database" / "acoustic_production_v2.db"
DEFAULT_OUTPUT_DIRECTORY = SCRIPT_DIRECTORY / "reports" / "rawdata_excel"

NAVY = "12304A"
BLUE = "2563EB"
LIGHT_BLUE = "E8F1FB"
WHITE = "FFFFFF"
GREEN_FILL = "DCFCE7"
GREEN_TEXT = "166534"
RED_FILL = "FEE2E2"
RED_TEXT = "B91C1C"
GRID = "D9E2EA"


def _resolve_path(value: str, base: Path) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if path.is_absolute() else (base / path).resolve()


def _safe_name(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value.strip())
    return cleaned.strip(" ._") or "measurement"


def _display_header(value: Any) -> Any:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return str(int(number)) if number.is_integer() else str(number)


def _write_dataframe(worksheet: Any, frame: Any) -> None:
    headers = [_display_header(column) for column in frame.columns]
    worksheet.append(headers)
    for row in frame.itertuples(index=False, name=None):
        worksheet.append(list(row))


def _style_data_sheet(worksheet: Any, row_count: int, column_count: int) -> None:
    worksheet.freeze_panes = "F2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 26
    thin = Side(style="thin", color=GRID)
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    widths = {"A": 22, "B": 18, "C": 13, "D": 12, "E": 11}
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width
    for column_index in range(6, column_count + 1):
        letter = worksheet.cell(1, column_index).column_letter
        worksheet.column_dimensions[letter].width = 10

    first_value_column = 10 if worksheet.cell(1, 6).value == "AttemptNo" else 6
    for row_index in range(2, row_count + 2):
        worksheet.cell(row_index, 3).number_format = "yyyy-mm-dd"
        worksheet.cell(row_index, 4).number_format = "hh:mm:ss"
        for column_index in range(first_value_column, column_count + 1):
            worksheet.cell(row_index, column_index).number_format = "0.000"

    if row_count:
        last_cell = worksheet.cell(row_count + 1, column_count).coordinate
        table = Table(displayName="RawDataTable", ref="A1:{}".format(last_cell))
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        status_range = "E2:E{}".format(row_count + 1)
        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=['UPPER($E2)="PASS"'],
                fill=PatternFill("solid", fgColor=GREEN_FILL),
                font=Font(color=GREEN_TEXT, bold=True),
            ),
        )
        worksheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=['UPPER($E2)="FAIL"'],
                fill=PatternFill("solid", fgColor=RED_FILL),
                font=Font(color=RED_TEXT, bold=True),
            ),
        )


def _write_metadata_sheet(
    workbook: Workbook,
    database_path: Path,
    test_date: str,
    measurement_name: str,
    raw_name: str,
    row_count: int,
    dataset_mode: str,
    projects: Sequence[str],
    station: str,
    specs: Any,
) -> None:
    worksheet = workbook.create_sheet("Metadata")
    worksheet.sheet_view.showGridLines = False
    rows = [
        ("Item", "Value"),
        ("GeneratedAt", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Database", str(database_path)),
        ("TestDate", test_date),
        ("Measurement", measurement_name),
        ("RawMeasurementName", raw_name),
        ("DatasetMode", "FIRST_VALID_PER_SN" if dataset_mode == "first" else "ALL_ATTEMPTS"),
        ("Projects", ", ".join(projects) if projects else "ALL"),
        ("Station", station),
        ("ExportedRows", row_count),
    ]
    for row in rows:
        worksheet.append(row)
    worksheet.append(())
    worksheet.append(
        (
            "project_code",
            "measurement_name",
            "value_unit",
            "weighting",
            "target",
            "lsl",
            "usl",
            "spec_revision",
            "spec_status",
        )
    )
    selected = specs.loc[specs["measurement_name"] == measurement_name] if not specs.empty else specs
    for row in selected.itertuples(index=False):
        worksheet.append(
            (
                getattr(row, "project_code"),
                getattr(row, "measurement_name"),
                getattr(row, "value_unit"),
                getattr(row, "weighting"),
                getattr(row, "target"),
                getattr(row, "lsl"),
                getattr(row, "usl"),
                getattr(row, "spec_revision"),
                getattr(row, "spec_status"),
            )
        )
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    header_row = 12
    for cell in worksheet[header_row]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 72
    for column_index in range(3, 10):
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = 18


def export_one_workbook(
    database_path: Path,
    output_path: Path,
    test_date: str,
    measurement_name: str,
    projects: Sequence[str],
    station: str,
    sn: Optional[str],
    dataset_mode: str,
    include_audit_columns: bool,
) -> Tuple[Path, int]:
    connection = connect_readonly(database_path)
    try:
        frame = build_export_frame(
            connection,
            measurement_name=measurement_name,
            test_date=test_date,
            projects=projects,
            station=station,
            sn=sn,
            dataset_mode=dataset_mode,
            include_audit_columns=include_audit_columns,
        )
        specs = query_current_specs(connection, projects, station)
    finally:
        connection.close()
    if frame.empty:
        raise ValueError(
            "{} 的 {} 沒有符合資料".format(test_date, measurement_name)
        )

    raw_name = raw_measurement_name(measurement_name)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_name(raw_name)[:31]
    _write_dataframe(worksheet, frame)
    _style_data_sheet(worksheet, len(frame), len(frame.columns))
    _write_metadata_sheet(
        workbook,
        database_path,
        test_date,
        measurement_name,
        raw_name,
        len(frame),
        dataset_mode,
        projects,
        station,
        specs,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(output_path.stem + ".tmp.xlsx")
    workbook.save(str(temporary_path))
    temporary_path.replace(output_path)
    return output_path, len(frame)


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="從聲學 SQLite 依日期與測項產生 RawData Excel。"
    )
    parser.add_argument("--database", default=str(DEFAULT_DATABASE))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIRECTORY))
    parser.add_argument(
        "--project",
        action="append",
        default=[],
        help="限定 project_code；可重複指定。",
    )
    parser.add_argument("--station", default="ST01")
    parser.add_argument("--sn", help="限定單一 SN。")
    parser.add_argument("--date", help="限定單日 YYYY-MM-DD。")
    parser.add_argument("--date-from", help="起始日期 YYYY-MM-DD。")
    parser.add_argument("--date-to", help="結束日期 YYYY-MM-DD。")
    parser.add_argument(
        "--measurement",
        action="append",
        default=[],
        help="測項名稱；可重複指定，例如 FREQRESP_IN_1。",
    )
    parser.add_argument(
        "--dataset",
        choices=("all", "first"),
        default="all",
        help="all 包含重測；first 每個 SN 只取第一次有效測試。",
    )
    parser.add_argument(
        "--audit-columns",
        action="store_true",
        help="加入 AttemptNo、IsRetest、SourceFile、TestRunID。",
    )
    return parser


def run(arguments: argparse.Namespace) -> List[Path]:
    database_path = _resolve_path(arguments.database, SCRIPT_DIRECTORY)
    output_directory = _resolve_path(arguments.output_dir, SCRIPT_DIRECTORY)
    date_from = arguments.date_from
    date_to = arguments.date_to
    if arguments.date:
        selected_date = validate_iso_date(arguments.date, "--date")
        date_from = selected_date
        date_to = selected_date
    measurements = [
        normalize_measurement_name(value) for value in arguments.measurement
    ]
    connection = connect_readonly(database_path)
    try:
        groups = list_export_groups(
            connection,
            projects=arguments.project,
            station=arguments.station,
            sn=arguments.sn,
            date_from=date_from,
            date_to=date_to,
            dataset_mode=arguments.dataset,
            measurement_names=measurements,
        )
    finally:
        connection.close()
    if groups.empty:
        raise ValueError("目前篩選條件沒有可匯出的 RawData")

    unique_groups = groups.drop_duplicates(["test_date", "measurement_name"])
    outputs = []  # type: List[Path]
    for row in unique_groups.itertuples(index=False):
        test_date = str(row.test_date)
        measurement_name = str(row.measurement_name)
        filename = "{}_{}.xlsx".format(
            test_date.replace("-", ""),
            _safe_name(raw_measurement_name(measurement_name)),
        )
        path, row_count = export_one_workbook(
            database_path=database_path,
            output_path=output_directory / filename,
            test_date=test_date,
            measurement_name=measurement_name,
            projects=arguments.project,
            station=arguments.station,
            sn=arguments.sn,
            dataset_mode=arguments.dataset,
            include_audit_columns=arguments.audit_columns,
        )
        print("[OK] {}（{} 列）".format(path, row_count))
        outputs.append(path)
    return outputs


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_argument_parser()
    arguments = parser.parse_args(argv)
    try:
        run(arguments)
        return 0
    except (FileNotFoundError, ValueError, sqlite3.Error, ImportError) as error:
        print("Excel 匯出失敗：{}".format(error), file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
